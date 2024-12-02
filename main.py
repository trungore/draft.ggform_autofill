import requests
import random 
import openpyxl
import time
from datetime import datetime
    
def str_time_prop(start, end, time_format, prop):
    """Get a time at a proportion of a range of two formatted times.

    start and end should be strings specifying times formatted in the
    given format (strftime-style), giving an interval [start, end].
    prop specifies how a proportion of the interval to be taken after
    start.  The returned time will be in the specified format.
    """

    stime = time.mktime(time.strptime(start, time_format))
    etime = time.mktime(time.strptime(end, time_format))

    ptime = stime + prop * (etime - stime)

    return time.strftime(time_format, time.localtime(ptime))


def random_date(start, end):
    datetime_str = str_time_prop(start, end, '%m/%d/%Y %I:%M %p', random.random())
    return datetime.strptime(datetime_str, '%m/%d/%Y %I:%M %p')
    

# https://docs.google.com/forms/d/e/1FAIpQLSdTK5Xw15YX3v8KE91P0icNCXPQxieHbe0gU7dRfqdwTs8rvg/viewform?usp=pp_url&entry.1576564961=0123&entry.698922775=Người+bệnh&entry.1454705730=12&entry.1747679787=Nam&entry.147791392=1000-03-12&entry.1629100921=10:20&entry.1523988971=Khoa+Ngoại+tiết+niệu&entry.1213306052=Từ+3++-+7+ngày&entry.1732837110=1+Khoa&entry.524248420=Tốt&entry.1105440216=Có+hướng+dẫn+tư+vấn+giáo+dục+sức+khỏe
# &entry.1576564961=0123
# &entry.698922775=Người+bệnh
# &entry.1454705730=12
# &entry.1747679787=Nam
# &entry.147791392=1000-03-12
# &entry.1629100921=10:20
# &entry.1523988971=Khoa+Ngoại+tiết+niệu
# &entry.1213306052=Từ+3++-+7+ngày
# &entry.1732837110=1+Khoa&entry.524248420=Tốt
# &entry.1105440216=Có+hướng+dẫn+tư+vấn+giáo+dục+sức+khỏe

# ==== SETUP EXCEL FILE ====
excel_path = "./data2.xlsx"
wb_obj = openpyxl.load_workbook(excel_path)
sheet_obj = wb_obj.active

# ==== SETUP GG FORM ====
prefix_url = "https://docs.google.com/forms/d/e/1FAIpQLSdTK5Xw15YX3v8KE91P0icNCXPQxieHbe0gU7dRfqdwTs8rvg/formResponse?usp=pp_url"
entryIds = [
    "UNUSED", # 0
    "1576564961", # 1
    "698922775", # 2
    "1454705730", # 3
    "1747679787", # 4
    "147791392", # 5
    "1629100921", # 6
    "1523988971", # 7
    "1213306052", # 8
    "1732837110", # 9
    "524248420", # 10
    "1105440216", # 11

    "285958344", # 11.1 - 1
    "285958344", # 11.1 - 2
    "285958344", # 11.1 - 3
    "285958344", # 11.1 - 4
    "285958344", # 11.1 - 5
    "285958344", # 11.1 - 6
    "285958344", # 11.1 - 7

    "1011649990", # 11.2 - 1
    "1011649990", # 11.2 - 2
    "1011649990", # 11.2 - 3

    "1380223014", # 11.3

    "936376113", # 11.4 - 1
    "936376113", # 11.4 - 2
    "936376113", # 11.4 - 3

    "1184618096", # 11.5 - 1
    "1638633367", # 11.5 - 2
    "2046749282", # 11.5 - 3
    "793872239", # 11.5 - 4

    "58084608", # 11.6
    "684735009" # 11.7
]
answers = [
    [], # 0. UNUSED
    [], # 1. SDT (excel)
    ["Người+bệnh", "Người+nhà+chăm+nuôi+người+bệnh"], # 2. Doi tuong khao sat 
    [], # 3. Tuoi (excel)
    [], # 4. Gioi tinh (excel)
    [], # 5. Ngay khao sat (random)
    [], # 6. Gio khao sat (random)

    ["Khoa+Ngoại+tiết+niệu"], # 7. Khoa dieu tri hien tai 
    ["Dưới+3+ngày", "Từ+3++-+7+ngày", "Từ+7+-+10+ngày", "Trên+10+ngày"], # 8. Thoi gian nam vien 
    ["1+Khoa"], # 9. So khoa dieu tri 
    ["Tốt", "Không+có+góp+ý+gì+thêm"], # 10. Gop y 
    ["Có+hướng+dẫn+tư+vấn+giáo+dục+sức+khỏe"], # 11. Huong dan 

    # 11.1. Nội dung hướng dẫn, tư vấn của nhân viên y tế cho người bệnh
    ["Nội+quy,+quy+định+của+bệnh+viện"], # 11.1 - 1. 
    ["Các+dịch+vụ+hiện+có+của+bệnh+viện"], # 11.1 - 2. 
    ["Chế+độ,+chính+sách+khi+người+bệnh+vào+viện"], # 11.1 - 3. 
    ["Chăm+sóc,+theo+dõi+người+bệnh"], # 11.1 - 4. 
    ["Chế+độ+ăn+cho+người+bệnh"], # 11.1 - 5. 
    ["Phục+hồi+chức+năng+cho+người+bệnh"], # 11.1 - 6. 
    ["Dự+phòng+về+bệnh,+biến+chứng+xảy+ra+liên+quan+đến+bệnh"], # 11.1 - 7. 

    # 11.2. Thời điểm hướng dẫn tư vấn cho người bệnh
    ["Khi+người+bệnh+vào+viện"], # 11.2 - 1.
    ["Trong+thời+gian+nằm+viện"], # 11.2 - 2.
    ["Trước+khi+người+bệnh+ra+viện"], # 11.2 - 3.

    ["Hàng+ngày+trong+đợt+điều+trị"], # 11.3. Số lần tư vấn trong đợt điều trị

    # 11.4. Người bệnh Tiếp cận nội dung hướng dẫn Giáo dục sức khỏe từ
    ["Nhân+viên+y+tế+tư+vấn+trực+tiếp+cho+người+bệnh/+người+nhà"], # 11.4 - 1.
    ["Quan+thông+tin+bằng+tranh+ảnh/+góc+truyền+thông"], # 11.4 - 2.
    ["Qua+truyền+tai+nhau+từ+người+nhà"], # 11.4 - 3.

    # 11.5. Người bệnh/ Người nhà có kiến thức để tự theo dõi, chăm sóc
    ["Có+kiến+thức"], # 11.5 - 1 
    ["Có+kiến+thức"], # 11.5 - 2 
    ["Có+kiến+thức"], # 11.5 - 3 
    ["Có+kiến+thức"], # 11.5 - 4

    ["5"], # 11.6. Mức độ tự tin tưởng của người bệnh
    ["5"], # 11.7. Mức độ hài lòng của người bệnh
]

# ==== LOCAL VAR ====
done = {}
start_index = 10
finish_index = 204
result_file = open("result.txt", "w")

# ==== EXCEL COLUMN ID ====
phone_number_col_id = 8
birth_year_col_id = 7 
sex_col_id = 6

def main(): 
    cnt = 0

    for row_id in range(start_index, finish_index + 1):
        if cnt == 195: 
            break

        if done.get(row_id) is not None: 
            continue

        url = prefix_url

        # 1. SDT (excel)
        if True: 
            phone_number = sheet_obj.cell(row = row_id, column = phone_number_col_id).value
            if phone_number is None:
                continue
            url += "&entry." + entryIds[1] + "=" + phone_number 

        # 2. Doi tuong khao sat (random)
        if True: 
            randId = random.randint(0, len(answers[2]) - 1)
            ans = answers[2][randId]
            url += "&entry." + entryIds[2] + "=" + ans 
        
        # 3. Tuoi (excel)     
        if True: 
            birth_year = sheet_obj.cell(row = row_id, column = birth_year_col_id).value
            if birth_year is None:
                continue
            age = 2024 - int(birth_year)
            url += "&entry." + entryIds[3] + "=" + ans 
        
        # 4. Gioi tinh (excel)   
        if True: 
            sex = sheet_obj.cell(row = row_id, column = sex_col_id).value
            if sex is None:
                continue
            url += "&entry." + entryIds[4] + "=" + sex 

        # 5. Ngay khao sat (random)
        # 6. Gio khao sat (random)
        if True: 
            datetime = random_date("8/20/2024 6:00 AM", "8/29/2024 9:00 PM")

            date = str(datetime.date())
            url += "&entry." + entryIds[5] + "=" + date 
            
            # hour_minute = str(datetime.hour) + ":" + str(datetime.minute) 
            # url += "&entry." + entryIds[6] + "=" + hour_minute 
            url += "&entry." + entryIds[6] + "_hour=" + str(datetime.hour)
            url += "&entry." + entryIds[6] + "_minute=" + str(datetime.minute)

        # 7. Khoa dieu tri hien tai 
        # 8. Thoi gian nam vien 
        # 9. So khoa dieu tri 
        # 10. Gop y
        # 11. Huong dan
        # ...
        for id in range(7, 7 + 25):       
            randId = random.randint(0, len(answers[id]) - 1)
            ans = answers[id][randId]
            url += "&entry." + entryIds[id] + "=" + ans 

        res = requests.post(url, timeout=10)  
        if res.status_code == 200:
            cnt += 1
            print(str(row_id) + " --> " + str(res))
            done[row_id] = True
        else: 
            print("FAIL: " + str(row_id))

        time.sleep(1)

    result_file.write(str(cnt) + "\n")

# -----------------------------------
if __name__ == "__main__":
    with open('log.txt', 'r') as log_file_reader:
        for line in log_file_reader:
            row_id = int( line.strip() )
            done[row_id] = True
        log_file_reader.close()

    log_file = open("log.txt", "w")

    for try_id in range(0, 8):
        main()

    for row_id in range(start_index, finish_index):
        if done.get(row_id) is None:
            log_file.write(str(row_id) + "\n")

    log_file.close()
    result_file.close()